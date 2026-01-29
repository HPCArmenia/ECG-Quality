import os
import numpy as np
import pandas as pd
import wfdb
import heartpy as hp
import warnings
from datetime import datetime
from scipy.signal import butter, filtfilt
from scipy.stats import kurtosis
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print("Current Time =", current_time)

# ==================================================
# SILENCE EXPECTED WARNINGS
# ==================================================
warnings.filterwarnings(
    "ignore",
    message="Precision loss occurred in moment calculation"
)

# ==================================================
# PATHS
# ==================================================
ROOT_WAVES = r"D:/Docs/NAS RA iiap/pHD/2025/Macedonia/datalake/raw/physionet.org/files/mimic4wdb/0.1.0/waves"
METADATA = r"mimic4wdb_metadata_step1.csv"
OUTPUT_XLSX = r"mimic4wdb_quality_results.xlsx"

# ==================================================
# FILTER (FOR METRICS ONLY — NO CLEANING)
# ==================================================
def bandpass(signal, fs, low=0.5, high=35, order=4):
    nyq = 0.5 * fs
    high = min(high, 0.99 * nyq)
    if low >= high or len(signal) < 10:
        return signal
    b, a = butter(order, [low / nyq, high / nyq], btype="band")
    return filtfilt(b, a, signal)

# ==================================================
# QUALITY METRICS (STEP 3)
# ==================================================
def detect_flat_metrics(signal, fs, window_sec=1.0):
    window = int(window_sec * fs)
    if len(signal) < window:
        return 0.0, 0.0
    diff = np.diff(signal)
    flat_mask = diff == 0
    return np.sum(flat_mask) / fs, np.sum(flat_mask) / len(signal)

def compute_snr(signal, fs):
    try:
        clean = bandpass(signal, fs)
        noise = signal - clean
        if np.var(noise) == 0:
            return np.nan
        return 10 * np.log10(np.var(clean) / np.var(noise))
    except:
        return np.nan

def baseline_wander(signal, fs):
    try:
        win = int(0.8 * fs)
        baseline = np.convolve(signal, np.ones(win)/win, mode="same")
        return np.std(baseline)
    except:
        return np.nan

def hf_noise(signal, fs):
    try:
        low = bandpass(signal, fs, low=0.5, high=15)
        return np.std(signal - low)
    except:
        return np.nan

def compute_kurtosis(signal):
    try:
        return kurtosis(signal, fisher=True, bias=False)
    except:
        return np.nan

def missing_percentage(signal):
    return np.sum(np.isnan(signal)) / len(signal)

# ==================================================
# HR COMPUTATION (STEP 5)
# ==================================================
def compute_hr(signal, fs):
    try:
        _, m = hp.process(signal, fs)
        return m["bpm"]
    except:
        return np.nan

def rpeak_metrics(signal, fs):
    try:
        wd, m = hp.process(signal, fs)
        rr = np.diff(wd['peaklist']) / fs
        return {
            "r_peak_detectable": True,
            "n_peaks": len(wd['peaklist']),
            "rr_std": np.std(rr)
        }
    except:
        return {
            "r_peak_detectable": False,
            "n_peaks": 0,
            "rr_std": np.nan
        }

# ==================================================
# WINDOW HR (STEP 6)
# ==================================================
def window_hr(signal, fs, window_sec=10):
    window = int(window_sec * fs)
    hrs = []
    for i in range(0, len(signal) - window, window):
        hr = compute_hr(signal[i:i+window], fs)
        if not np.isnan(hr):
            hrs.append(hr)
    return np.mean(hrs) if hrs else np.nan

# ==================================================
# MAIN LOOP
# ==================================================
df = pd.read_csv(METADATA)
results = []

for _, row in df.iterrows():

    subject = row["subject_id"]
    study = row["study_id"]
    record_id = row["record_id"]

    record_folder = os.path.join(
        ROOT_WAVES, subject, study, record_id.split("_")[0]
    )
    rec_path = os.path.join(record_folder, record_id)

    if not os.path.exists(rec_path + ".hea"):
        continue

    try:
        rec = wfdb.rdrecord(rec_path)
    except:
        continue

    fs = rec.fs  # WFDB logic (defaults to 250 Hz if missing)

    signal_names = [s.upper().strip() for s in rec.sig_name]
    if "II" not in signal_names:
        continue

    ecg = rec.p_signal[:, signal_names.index("II")]

    flat_sec, flat_pct = detect_flat_metrics(ecg, fs)
    rp = rpeak_metrics(ecg, fs)
    results.append({
        "subject_id": subject,
        "study_id": study,
        "record_id": record_id,
        "fs_hz": fs,
        "signal_length": len(ecg),
        "flat_seconds": round(flat_sec, 3),
        "flat_percentage": round(flat_pct, 3),
        "missing_percentage": round(missing_percentage(ecg), 3),
        "snr": compute_snr(ecg, fs),
        "baseline_wander": baseline_wander(ecg, fs),
        "hf_noise": hf_noise(ecg, fs),
        "kurtosis": compute_kurtosis(ecg),
        "estimated_hr": compute_hr(ecg, fs),   # STEP 5
        "window_hr": window_hr(ecg, fs, 10),    # STEP 6
        "r_peak_detectable": rp["r_peak_detectable"],
        "n_r_peaks": rp["n_peaks"],
        "rr_std": rp["rr_std"]
    })

# ==================================================
# SAVE TO EXCEL
# ==================================================
quality_df = pd.DataFrame(results)
quality_df.to_excel(OUTPUT_XLSX, index=False)

# ==================================================
# CONDITIONAL FORMATTING (EXCEL-NATIVE)
# ==================================================
wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

cols = {c.value: c.column_letter for c in ws[1]}
last_row = ws.max_row

# ---------- Flatline ----------
col = cols["flat_percentage"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    CellIsRule(
        operator="greaterThan",
        formula=["0.1"],
        fill=red
    )
)

# ---------- Missing data ----------
col = cols["missing_percentage"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    CellIsRule(
        operator="greaterThan",
        formula=["0.1"],
        fill=red
    )
)

# ---------- SNR ----------
col = cols["snr"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    CellIsRule(
        operator="lessThan",
        formula=["5"],
        fill=red
    )
)

# ---------- Baseline wander ----------
col = cols["baseline_wander"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    CellIsRule(
        operator="greaterThan",
        formula=["0.5"],
        fill=red
    )
)

# ---------- HR failure ----------
col = cols["estimated_hr"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    FormulaRule(
        formula=[
            f"OR({col}2<40,{col}2>180,ISBLANK({col}2))"
        ],
        fill=red
    )
)

# ---------- kurtosis failure ----------
col = cols["kurtosis"]
ws.conditional_formatting.add(
    f"{col}2:{col}{last_row}",
    FormulaRule(
        formula=[
            f"OR({col}2<5,{col}2>30,ISBLANK({col}2))"
        ],
        fill=red
    )
)
wb.save(OUTPUT_XLSX)


now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print("Current Time =", current_time)
